import io
import types
from unittest.mock import patch

import pytest
from httpx import AsyncClient


@pytest.fixture
async def project_with_prices(auth_client: AsyncClient) -> str:
    """Create project, upload smeta, set contractor price."""
    create_resp = await auth_client.post(
        "/api/v1/projects",
        json={"name": "Margin Test", "description": ""},
    )
    project_id = create_resp.json()["id"]

    from openpyxl import Workbook
    buf = io.BytesIO()
    wb = Workbook()
    wb.active.append(["header"])
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    mock_item = types.SimpleNamespace(
        number=1, code="ГЭСН11-01-009-04", name="Штукатурка стен",
        unit="м²", quantity=120.0, unit_price=700.0, total_price=84000.0,
        item_type="work", section="",
    )
    mock_result = types.SimpleNamespace(items=[mock_item])

    with (
        patch("app.services.smeta_service.parse_excel", return_value=mock_result),
        patch("app.services.smeta_service.aggregate_materials", return_value=[]),
    ):
        await auth_client.post(
            f"/api/v1/projects/{project_id}/smeta/upload",
            files={
                "file": (
                    "smeta.xlsx",
                    io.BytesIO(xlsx_bytes),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    items_resp = await auth_client.get(
        f"/api/v1/projects/{project_id}/smeta/items"
    )
    smeta_item_id = items_resp.json()["items"][0]["id"]

    await auth_client.put(
        f"/api/v1/projects/{project_id}/contractor-prices/{smeta_item_id}",
        json={"price": 580.0},
    )
    return project_id


@pytest.mark.asyncio
async def test_margin_not_available_without_smeta(
    auth_client: AsyncClient
) -> None:
    resp = await auth_client.post(
        "/api/v1/projects", json={"name": "Empty", "description": ""}
    )
    project_id = resp.json()["id"]

    resp = await auth_client.get(f"/api/v1/projects/{project_id}/margin")
    assert resp.status_code == 400


@pytest.mark.asyncio
async def test_margin_calculation(
    auth_client: AsyncClient, project_with_prices: str
) -> None:
    resp = await auth_client.get(f"/api/v1/projects/{project_with_prices}/margin")
    assert resp.status_code == 200
    data = resp.json()

    # ceiling = 84000, cost = 580 * 120 = 69600, margin = 14400
    assert data["total_ceiling"] == pytest.approx(84000.0)
    assert data["total_cost"] == pytest.approx(69600.0)
    assert data["total_margin"] == pytest.approx(14400.0)
    assert data["margin_pct"] == pytest.approx(17.14, rel=0.01)
    assert data["min_profit"] == pytest.approx(4200.0)
    assert data["floor_price"] == pytest.approx(79800.0)

    assert len(data["items"]) == 1
    item = data["items"][0]
    assert item["name"] == "Штукатурка стен"
    assert item["ceiling_price"] == pytest.approx(84000.0)
    assert item["cost_price"] == pytest.approx(69600.0)
    assert item["margin_pct"] == pytest.approx(17.14, rel=0.01)
    assert item["status"] == "green"


@pytest.mark.asyncio
async def test_margin_item_status_yellow(auth_client: AsyncClient) -> None:
    create_resp = await auth_client.post(
        "/api/v1/projects", json={"name": "Yellow Margin", "description": ""}
    )
    project_id = create_resp.json()["id"]

    from openpyxl import Workbook
    buf = io.BytesIO()
    wb = Workbook()
    wb.active.append(["h"])
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    mock_item = types.SimpleNamespace(
        number=1, code="CODE-01", name="Item",
        unit="шт", quantity=1.0, unit_price=100.0,
        total_price=100.0, item_type="work", section="",
    )
    mock_result = types.SimpleNamespace(items=[mock_item])

    with (
        patch("app.services.smeta_service.parse_excel", return_value=mock_result),
        patch("app.services.smeta_service.aggregate_materials", return_value=[]),
    ):
        await auth_client.post(
            f"/api/v1/projects/{project_id}/smeta/upload",
            files={"file": ("s.xlsx", io.BytesIO(xlsx_bytes), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    items_resp = await auth_client.get(f"/api/v1/projects/{project_id}/smeta/items")
    smeta_item_id = items_resp.json()["items"][0]["id"]
    await auth_client.put(
        f"/api/v1/projects/{project_id}/contractor-prices/{smeta_item_id}",
        json={"price": 90.0},
    )

    resp = await auth_client.get(f"/api/v1/projects/{project_id}/margin")
    assert resp.status_code == 200
    assert resp.json()["items"][0]["status"] == "yellow"


@pytest.mark.asyncio
async def test_margin_item_status_loss(auth_client: AsyncClient) -> None:
    create_resp = await auth_client.post(
        "/api/v1/projects", json={"name": "Loss Project", "description": ""}
    )
    project_id = create_resp.json()["id"]

    from openpyxl import Workbook
    buf = io.BytesIO()
    wb = Workbook()
    wb.active.append(["h"])
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    mock_item = types.SimpleNamespace(
        number=1, code="CODE-02", name="Loss Item",
        unit="шт", quantity=1.0, unit_price=100.0,
        total_price=100.0, item_type="work", section="",
    )
    mock_result = types.SimpleNamespace(items=[mock_item])

    with (
        patch("app.services.smeta_service.parse_excel", return_value=mock_result),
        patch("app.services.smeta_service.aggregate_materials", return_value=[]),
    ):
        await auth_client.post(
            f"/api/v1/projects/{project_id}/smeta/upload",
            files={"file": ("s.xlsx", io.BytesIO(xlsx_bytes), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    items_resp = await auth_client.get(f"/api/v1/projects/{project_id}/smeta/items")
    smeta_item_id = items_resp.json()["items"][0]["id"]
    await auth_client.put(
        f"/api/v1/projects/{project_id}/contractor-prices/{smeta_item_id}",
        json={"price": 120.0},
    )

    resp = await auth_client.get(f"/api/v1/projects/{project_id}/margin")
    assert resp.status_code == 200
    assert resp.json()["items"][0]["status"] == "loss"


@pytest.mark.asyncio
async def test_margin_export(
    auth_client: AsyncClient, project_with_prices: str
) -> None:
    resp = await auth_client.post(
        f"/api/v1/projects/{project_with_prices}/margin/export"
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert len(resp.content) > 0


@pytest.mark.asyncio
async def test_margin_export_content(
    auth_client: AsyncClient, project_with_prices: str
) -> None:
    resp = await auth_client.post(
        f"/api/v1/projects/{project_with_prices}/margin/export"
    )
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert ws.cell(1, 1).value == "Итого НМЦК (потолок)"
    assert ws.cell(1, 2).value == pytest.approx(84000.0)
