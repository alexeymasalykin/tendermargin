from __future__ import annotations

import io
import uuid
from typing import List

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.contractor import ContractorPrice
from app.models.smeta import SmetaItem
from app.schemas.margin import MarginItem, MarginResult

MIN_PROFIT_PCT = 0.05


def _item_status(margin_pct: float) -> str:
    if margin_pct < 0:
        return "loss"
    if margin_pct < 5:
        return "red"
    if margin_pct < 15:
        return "yellow"
    return "green"


async def calculate_margin(
    project_id: uuid.UUID, db: AsyncSession
) -> MarginResult:
    result = await db.execute(
        select(SmetaItem, ContractorPrice)
        .outerjoin(
            ContractorPrice,
            (ContractorPrice.smeta_item_id == SmetaItem.id)
            & (ContractorPrice.project_id == project_id),
        )
        .where(SmetaItem.project_id == project_id)
        .order_by(SmetaItem.number)
    )
    rows = result.all()

    if not rows:
        raise HTTPException(
            status_code=400,
            detail="No smeta data found. Upload a smeta first.",
        )

    has_any_price = any(cp is not None and cp.price is not None for _, cp in rows)
    if not has_any_price:
        raise HTTPException(
            status_code=400,
            detail="No contractor prices set. Fill at least one contractor price.",
        )

    items: List[MarginItem] = []
    total_ceiling = 0.0
    total_cost = 0.0

    for si, cp in rows:
        ceiling = float(si.total_price)
        if cp is not None and cp.price is not None:
            cost = float(cp.price) * float(si.quantity)
        else:
            cost = ceiling

        margin = ceiling - cost
        margin_pct = (margin / ceiling * 100) if ceiling > 0 else 0.0

        items.append(
            MarginItem(
                name=si.name,
                code=si.code or "",
                item_type=si.item_type or "unknown",
                unit=si.unit or "",
                quantity=float(si.quantity),
                ceiling_price=ceiling,
                cost_price=cost,
                margin=round(margin, 2),
                margin_pct=round(margin_pct, 2),
                status=_item_status(margin_pct),
            )
        )
        total_ceiling += ceiling
        total_cost += cost

    total_margin = total_ceiling - total_cost
    margin_pct_total = (
        (total_margin / total_ceiling * 100) if total_ceiling > 0 else 0.0
    )
    min_profit = total_ceiling * MIN_PROFIT_PCT
    floor_price = total_ceiling - min_profit
    max_discount_pct = (
        ((total_margin - min_profit) / total_ceiling * 100)
        if total_ceiling > 0
        else 0.0
    )

    return MarginResult(
        total_ceiling=round(total_ceiling, 2),
        total_cost=round(total_cost, 2),
        total_margin=round(total_margin, 2),
        margin_pct=round(margin_pct_total, 2),
        min_profit=round(min_profit, 2),
        max_discount_pct=round(max_discount_pct, 2),
        floor_price=round(floor_price, 2),
        items=items,
    )


async def export_margin_excel(
    project_id: uuid.UUID, db: AsyncSession
) -> bytes:
    result = await calculate_margin(project_id, db)

    wb = Workbook()
    ws = wb.active
    ws.title = "Расчёт маржи"

    ws.append(["Итого НМЦК (потолок)", result.total_ceiling])
    ws.append(["Себестоимость", result.total_cost])
    ws.append(["Маржа", result.total_margin])
    ws.append(["Маржа %", f"{result.margin_pct:.2f}%"])
    ws.append(["Мин. прибыль (5%)", result.min_profit])
    ws.append(["Макс. снижение %", f"{result.max_discount_pct:.2f}%"])
    ws.append(["Цена пол", result.floor_price])
    ws.append([])

    headers = [
        "№", "Наименование", "Код", "Тип", "Ед.изм.", "Кол-во",
        "Потолок", "Себестоимость", "Маржа", "Маржа %", "Статус",
    ]
    ws.append(headers)
    header_row_idx = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row_idx, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center")

    STATUS_COLORS = {
        "green": "16A34A",
        "yellow": "EAB308",
        "red": "DC2626",
        "loss": "7F1D1D",
    }
    STATUS_LABELS = {
        "green": "Норма",
        "yellow": "Внимание",
        "red": "Риск",
        "loss": "Убыток",
    }

    for idx, item in enumerate(result.items, start=1):
        ws.append([
            idx, item.name, item.code, item.item_type, item.unit,
            item.quantity, item.ceiling_price, item.cost_price,
            item.margin, f"{item.margin_pct:.2f}%",
            STATUS_LABELS.get(item.status, item.status),
        ])
        status_cell = ws.cell(row=ws.max_row, column=11)
        color = STATUS_COLORS.get(item.status, "64748B")
        status_cell.fill = PatternFill("solid", fgColor=color)
        status_cell.font = Font(color="FFFFFF")

    for col, width in zip("ABCDEFGHIJK", [5, 45, 18, 12, 10, 10, 15, 15, 15, 10, 12]):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
