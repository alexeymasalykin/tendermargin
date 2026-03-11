from __future__ import annotations

import io
import uuid
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.material import Material
from app.models.pricelist import PricelistMatch
from app.schemas.materials import MaterialOut


async def list_materials(
    project_id: uuid.UUID, db: AsyncSession
) -> List[MaterialOut]:
    result = await db.execute(
        select(Material, PricelistMatch)
        .outerjoin(
            PricelistMatch,
            (PricelistMatch.material_id == Material.id)
            & (PricelistMatch.project_id == project_id),
        )
        .where(Material.project_id == project_id)
        .order_by(Material.name)
    )
    rows = result.all()

    materials = []
    for mat, match in rows:
        sp = float(match.supplier_price) if match and match.supplier_price else None
        supplier_total = float(mat.quantity) * sp if sp is not None else None
        materials.append(
            MaterialOut(
                id=mat.id,
                name=mat.name,
                unit=mat.unit,
                quantity=float(mat.quantity),
                smeta_total=float(mat.smeta_total),
                supplier_price=sp,
                supplier_total=supplier_total,
            )
        )
    return materials


async def export_materials_excel(
    project_id: uuid.UUID, db: AsyncSession
) -> bytes:
    materials = await list_materials(project_id, db)

    wb = Workbook()
    ws = wb.active
    ws.title = "Материалы"

    header_fill = PatternFill("solid", fgColor="2563EB")
    headers = ["№", "Наименование", "Ед.изм.", "Объём", "Цена поставщика", "Сумма"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for idx, mat in enumerate(materials, start=1):
        ws.append([
            idx, mat.name, mat.unit, float(mat.quantity),
            mat.supplier_price, mat.supplier_total,
        ])

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
