"""LLM-based mapping of supplier pricelists to material list via OpenRouter."""

from __future__ import annotations

import json
import os
from dataclasses import dataclass
from typing import Optional

import openpyxl
import pdfplumber
from dotenv import load_dotenv
from openai import OpenAI

from core.materials import MaterialRow

load_dotenv()

MODEL = "anthropic/claude-sonnet-4"


def _get_client() -> OpenAI:
    api_key = os.getenv("OPENROUTER_API_KEY")
    if not api_key:
        raise ValueError("OPENROUTER_API_KEY not set in .env")
    return OpenAI(
        base_url="https://openrouter.ai/api/v1",
        api_key=api_key,
    )


def _chat(client: OpenAI, prompt: str, max_tokens: int = 2000) -> str:
    """Send a single prompt and return the text response."""
    response = client.chat.completions.create(
        model=MODEL,
        max_tokens=max_tokens,
        messages=[{"role": "user", "content": prompt}],
    )
    return response.choices[0].message.content.strip()


def _parse_json(text: str) -> dict | list:
    """Parse JSON from LLM response, stripping markdown fences if present."""
    if text.startswith("```"):
        text = text.split("\n", 1)[1].rsplit("```", 1)[0].strip()
    return json.loads(text)


@dataclass
class PricelistStructure:
    header_row: int
    name_col: int
    price_col: int
    unit_col: Optional[int]
    vat_included: bool
    vat_rate: float


@dataclass
class MappingMatch:
    material_index: int
    pricelist_index: Optional[int]
    confidence: float
    reason: str
    supplier_name: str = ""
    supplier_price: float = 0.0


def _is_pdf(source) -> bool:
    """Check if the source file is a PDF by extension or magic bytes."""
    path = str(source)
    return path.lower().endswith(".pdf")


def _read_pdf_rows(source, max_rows: int = 0) -> list[list]:
    """Extract table rows from a PDF file using pdfplumber."""
    all_rows: list[list] = []
    with pdfplumber.open(source) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                if table:
                    for row in table:
                        all_rows.append([str(v).strip() if v else "" for v in row])
                        if max_rows and len(all_rows) >= max_rows:
                            return all_rows
            # Fallback: if no tables found on page, try extracting text lines
            if not tables:
                text = page.extract_text() or ""
                for line in text.split("\n"):
                    parts = line.split()
                    if parts:
                        all_rows.append(parts)
                        if max_rows and len(all_rows) >= max_rows:
                            return all_rows
    return all_rows


def read_pricelist_preview(source, max_rows: int = 20) -> list[list]:
    """Read first N rows from supplier pricelist (Excel or PDF)."""
    if _is_pdf(source):
        return _read_pdf_rows(source, max_rows=max_rows)

    wb = openpyxl.load_workbook(source, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=min(max_rows, ws.max_row), values_only=True):
        rows.append([str(v) if v is not None else "" for v in row])
    wb.close()
    return rows


def _parse_price(val: str) -> float:
    """Parse a price string handling Russian formatting."""
    try:
        return float(val.replace(",", ".").replace(" ", "").replace("\xa0", ""))
    except (ValueError, TypeError):
        return 0.0


def read_pricelist_data(source, structure: PricelistStructure) -> list[dict]:
    """Read all items from supplier pricelist using detected structure (Excel or PDF)."""
    if _is_pdf(source):
        all_rows = _read_pdf_rows(source)
        rows_iter = all_rows[structure.header_row + 1:]
    else:
        wb = openpyxl.load_workbook(source, data_only=True)
        ws = wb.active
        rows_iter = list(ws.iter_rows(min_row=structure.header_row + 1, values_only=True))
        wb.close()

    items = []
    for row in rows_iter:
        row_list = list(row)
        name = str(row_list[structure.name_col]) if structure.name_col < len(row_list) and row_list[structure.name_col] else ""
        if not name.strip():
            continue

        price = 0.0
        if structure.price_col < len(row_list) and row_list[structure.price_col] is not None:
            price = _parse_price(str(row_list[structure.price_col]))

        unit = ""
        if structure.unit_col is not None and structure.unit_col < len(row_list):
            unit = str(row_list[structure.unit_col] or "")

        if price > 0:
            items.append({
                "index": len(items),
                "name": name.strip(),
                "price": price,
                "unit": unit.strip(),
            })

    return items


def detect_structure(source) -> PricelistStructure:
    """Use LLM to detect pricelist structure."""
    rows = read_pricelist_preview(source)
    rows_json = json.dumps(rows, ensure_ascii=False, indent=2)

    client = _get_client()
    text = _chat(client, f"""Ты анализируешь Excel-файл прайс-листа строительного поставщика.
Вот первые 20 строк файла (JSON):
{rows_json}

Определи:
1. Номер строки с заголовками (header_row, 0-based)
2. Номер колонки с наименованием товара (name_col, 0-based)
3. Номер колонки с ценой (price_col, 0-based)
4. Номер колонки с единицей измерения (unit_col, 0-based), если есть
5. Цена с НДС или без (vat_included: true/false)
6. Ставка НДС если указана (vat_rate)

Ответь строго в JSON без markdown:
{{"header_row": N, "name_col": N, "price_col": N, "unit_col": N, "vat_included": true, "vat_rate": 20}}""",
        max_tokens=500,
    )

    data = _parse_json(text)
    return PricelistStructure(
        header_row=data["header_row"],
        name_col=data["name_col"],
        price_col=data["price_col"],
        unit_col=data.get("unit_col"),
        vat_included=data.get("vat_included", True),
        vat_rate=data.get("vat_rate", 20),
    )


def map_materials(
    materials: list[MaterialRow],
    pricelist_items: list[dict],
    batch_size: int = 20,
) -> list[MappingMatch]:
    """Use LLM to match materials to pricelist items."""
    client = _get_client()
    all_matches: list[MappingMatch] = []

    materials_json = json.dumps(
        [{"index": m.index, "name": m.name, "unit": m.unit, "quantity": m.quantity}
         for m in materials],
        ensure_ascii=False,
    )

    # Process pricelist in batches
    for batch_start in range(0, len(pricelist_items), batch_size):
        batch = pricelist_items[batch_start:batch_start + batch_size]
        pricelist_json = json.dumps(batch, ensure_ascii=False)

        text = _chat(client, f"""Ты сопоставляешь позиции из прайс-листа поставщика с ведомостью материалов строительной сметы.

Ведомость материалов (что нужно найти):
{materials_json}

Позиции прайса поставщика:
{pricelist_json}

Для каждой позиции ведомости найди наиболее подходящую позицию из прайса.
Учитывай: марку, сечение, размеры, ГОСТ, тип. "ВВГнг(А)-LS 3x1,5" и "Кабель ВВГнг 3*1.5" — это одно и то же.

Ответь строго в JSON без markdown:
[
  {{
    "material_index": 0,
    "pricelist_index": 5,
    "confidence": 0.92,
    "reason": "Совпадение марки и сечения кабеля"
  }}
]

Если подходящей позиции нет, укажи "pricelist_index": null, "confidence": 0.""")

        matches = _parse_json(text)
        for match in matches:
            pl_idx = match.get("pricelist_index")
            supplier_name = ""
            supplier_price = 0.0
            if pl_idx is not None and 0 <= pl_idx < len(pricelist_items):
                supplier_name = pricelist_items[pl_idx]["name"]
                supplier_price = pricelist_items[pl_idx]["price"]

            all_matches.append(MappingMatch(
                material_index=match["material_index"],
                pricelist_index=pl_idx,
                confidence=match.get("confidence", 0),
                reason=match.get("reason", ""),
                supplier_name=supplier_name,
                supplier_price=supplier_price,
            ))

    return all_matches


def confidence_color(confidence: float) -> str:
    """Return color category for confidence level."""
    if confidence >= 0.85:
        return "green"
    if confidence >= 0.50:
        return "yellow"
    return "red"
